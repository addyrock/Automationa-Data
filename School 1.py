import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook

# Load the Excel file
file_path = 'C:\\Users\\arslan.arif\\Desktop\\Map Data\\schools.xlsx'  # Replace with your actual Excel file path
df = pd.read_excel(file_path)

# Create or load the output Excel file
output_file = "extracted_addresses.xlsx"
try:
    workbook = load_workbook(output_file)
    worksheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Location Name", "City Name", "Address", "Category", "Contact Number", "Latitude", "Longitude", "Google Name"])
    workbook.save(output_file)

# Set up Chrome options
chrome_options = Options()
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-software-rasterizer")

# Use ChromeDriverManager to install the correct driver version and set it up with Service

driver = webdriver.Chrome()

# Open Google Maps
driver.get("https://www.google.com/maps")
time.sleep(3)  # Wait for the page to load

# Iterate over each row in the Excel file
for index, row in df.iterrows():
    location_name = row['Location Name']
    city_name = row['City Name']

    try:
        # Find the search box and enter the location and city name
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "searchboxinput"))
        )
        search_box.clear()
        search_box.send_keys(f"{location_name}, {city_name}")
        search_box.send_keys(Keys.RETURN)

        # Wait for results to load
        time.sleep(5)

        # Extract the address
        try:
            address_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//*[contains(@class, 'Io6YTe') and contains(@class, 'fontBodyMedium')]")
                )
            )
            address = address_element.text.strip()
        except Exception:
            address = "Address not found"

        # Extract the category information
        try:
            category_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div/div[1]/div[2]/div/div[2]/span/span/button')
                )
            )
            category = category_element.text.strip()
        except Exception:
            category = "Category not found"

        # Extract the contact number
        try:
            contact_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[9]/div[6]/button/div/div[2]/div[1]')
                )
            )
            contact_element.click()
            contact_number = contact_element.text.strip()
        except Exception:
            contact_number = "Contact not found"

        # Extract coordinates from the URL
        try:
            current_url = driver.current_url
            if "/@" in current_url:
                coordinates_part = current_url.split("/@")[1].split(",")[:2]
                latitude = float(coordinates_part[0])
                longitude = float(coordinates_part[1])
            else:
                latitude = "Latitude not found"
                longitude = "Longitude not found"
        except Exception as coord_error:
            print(f"Error extracting coordinates for {location_name}, {city_name}: {str(coord_error)}")
            latitude = "Latitude not found"
            longitude = "Longitude not found"

        # Extract the name of the school (Google Name)
        try:
            google_name_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div/div[1]/div[1]/h1")
                )
            )
            google_name = google_name_element.text.strip()
        except Exception:
            google_name = "Google Name not found"

        print(f"Location: {location_name}, {city_name} - Address: {address}, Category: {category}, Contact: {contact_number}, Latitude: {latitude}, Longitude: {longitude}, Google Name: {google_name}")

        # Append data to the worksheet
        worksheet.append([location_name, city_name, address, category, contact_number, latitude, longitude, google_name])

    except Exception as e:
        print(f"Could not find information for {location_name}, {city_name}: {str(e)}")
        worksheet.append([location_name, city_name, "Address not found", "Category not found", "Contact not found", "Latitude not found", "Longitude not found", "Google Name not found"])

    # Save the workbook after each iteration
    workbook.save(output_file)

    # Wait before processing the next row
    time.sleep(2)

print("Data saved to extracted_addresses.xlsx.")

# Close the browser after processing all locations
driver.quit()
