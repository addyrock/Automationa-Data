import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook

# Load the Excel file
file_path = 'C:\\Users\\arslan.arif\\Desktop\\Map Data\\schools.xlsx'  # Replace with your actual Excel file path
df = pd.read_excel(file_path)

# Create or load the output Excel file
output_file = "extracted_addresses.xlsx"
try:
    workbook = load_workbook(output_file)
    worksheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Location Name", "City Name", "Address", "Category", "Contact Number", "Latitude", "Longitude", "Google Name"])
    workbook.save(output_file)

# Set up Chrome options
chrome_options = Options()
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-software-rasterizer")

# Use ChromeDriverManager to install the correct driver version and set it up with Service
driver = webdriver.Chrome()

# Open Google Maps
driver.get("https://www.google.com/maps")
time.sleep(3)  # Wait for the page to load

# Iterate over each row in the Excel file
for index, row in df.iterrows():
    location_name = row['Location Name']
    city_name = row['City Name']

    try:
        # Find the search box and enter the location and city name
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "searchboxinput"))
        )
        search_box.clear()
        search_box.send_keys(f"{location_name}, {city_name}")
        search_box.send_keys(Keys.RETURN)

        # Wait for results to load
        time.sleep(5)

        # Scroll to load all results
        scroll_attempts = 0
        while scroll_attempts < 5:
            driver.execute_script("window.scrollBy(0, 1000);")
            time.sleep(3)  # Wait for results to load
            scroll_attempts += 1

        # Extract all locations from the list of search results
        locations = driver.find_elements(By.CLASS_NAME, "section-result")
        for location in locations:
            # Extract the name and other details for each location
            try:
                name = location.find_element(By.CLASS_NAME, "section-result-title").text.strip()
            except:
                name = "Name not found"

            try:
                address = location.find_element(By.CLASS_NAME, "section-result-location").text.strip()
            except:
                address = "Address not found"

            try:
                contact = location.find_element(By.CLASS_NAME, "section-result-phone-number").text.strip()
            except:
                contact = "Contact not found"

            # Extract coordinates from the URL
            try:
                current_url = driver.current_url
                if "/@" in current_url:
                    coordinates_part = current_url.split("/@")[1].split(",")[:2]
                    latitude = float(coordinates_part[0])
                    longitude = float(coordinates_part[1])
                else:
                    latitude = "Latitude not found"
                    longitude = "Longitude not found"
            except Exception as coord_error:
                print(f"Error extracting coordinates for {location_name}, {city_name}: {str(coord_error)}")
                latitude = "Latitude not found"
                longitude = "Longitude not found"

            # Extract the Google name (if available)
            try:
                google_name = location.find_element(By.CLASS_NAME, "section-result-title").text.strip()
            except:
                google_name = "Google Name not found"

            # Print the extracted information
            print(f"Location: {name}, Address: {address}, Contact: {contact}, Latitude: {latitude}, Longitude: {longitude}, Google Name: {google_name}")

            # Append the extracted information to the Excel file
            worksheet.append([name, city_name, address, "Category not found", contact, latitude, longitude, google_name])

        # Save the workbook after processing the current location
        workbook.save(output_file)

    except Exception as e:
        print(f"Could not find information for {location_name}, {city_name}: {str(e)}")
        worksheet.append([location_name, city_name, "Address not found", "Category not found", "Contact not found", "Latitude not found", "Longitude not found", "Google Name not found"])

    # Wait before processing the next row
    time.sleep(2)

print("Data saved to extracted_addresses.xlsx.")

# Close the browser after processing all locations
driver.quit()
