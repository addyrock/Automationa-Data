import openpyxl

#receive data from excel sheet
# file="D:\\Arslan testing Data\\test data.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook["Sheet1"]
#
# rows=sheet.max_row
# cols=sheet.max_column
# #Reading all rows and colums frome excel sheet
# for r in range(1,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(r,c).value,end='     ')
#         print()

#send data into excel file
file="D:\\Arslan testing Data\\test data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="welcome"
workbook.save(file) #save the file


sheet.cell(1,1).value=567
sheet.cell(3,2).value="david"